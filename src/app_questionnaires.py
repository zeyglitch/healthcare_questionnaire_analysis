"""
Application GUI — Analyse de complétude des questionnaires hospitaliers.

Deux onglets :
  1. Configuration des questions obligatoires
  2. Contrôle Qualité – tirage aléatoire de réponses
"""

import os
import json
import sys
import random
from pathlib import Path
from tkinter import filedialog, messagebox
from datetime import datetime

import customtkinter as ctk

# ── Import des fonctions d'analyse ──
from analyse_questionnaires import (
    lire_csv,
    analyser_completude,
    generer_rapport_pdf,
    regrouper_par_service,
)

# ──────────────────────────────────────────────
#  Répertoire de l'application (compatible .exe)
# ──────────────────────────────────────────────
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CONFIG_FILE = os.path.join(APP_DIR, "application", "config_obligatoires.json")
if not os.path.exists(os.path.dirname(CONFIG_FILE)):
    CONFIG_FILE = "config_obligatoires.json"

# ──────────────────────────────────────────────
#  Palette de couleurs (Thème Midnight Multicolore)
# ──────────────────────────────────────────────
THEME = {
    "bg":              "#0F172A",     # Fond profond "Midnight Blue" (très reposant)
    "bg_sidebar":      "#1E293B",     # Sidebar légèrement plus claire
    "bg_card":         "#1E293B",     # Cartes au même niveau
    "bg_section":      "#334155",     # Gris bleuté pour sections internes
    
    # Boutons avec des couleurs distinctes et vibrantes
    "btn_load":        "#3B82F6",     # Bleu éclatant
    "btn_load_hover":  "#2563EB",     
    
    "btn_save":        "#8B5CF6",     # Violet vibrant
    "btn_save_hover":  "#7C3AED",     
    
    "btn_pdf":         "#F43F5E",     # Rose/Rouge dynamique
    "btn_pdf_hover":   "#E11D48",     

    "btn_excel":       "#10B981",     # Vert émeraude pour l'export Excel
    "btn_excel_hover": "#059669",
    
    # Éléments d'interface
    "header_card":     "#312E81",     # Indigo profond pour les en-têtes
    "checkbox":        "#0EA5E9",     # Cyan / Bleu ciel
    "checkbox_hover":  "#0284C7",     
    
    # Textes et bordures
    "text_title":      "#E0E7FF",     # Titres clairs (blanc bleuté)
    "text_main":       "#F8FAFC",     # Texte standard (blanc très doux)
    "text_dim":        "#94A3B8",     # Texte secondaire (gris)
    "text_light":      "#FFFFFF",     
    "border":          "#334155",     # Bordures subtiles
    
    # Statuts
    "success":         "#10B981",     
    "warning":         "#F59E0B",     
    "danger":          "#EF4444",     

    # Onglets
    "tab_selected":    "#3B82F6",
    "tab_unselected":  "#1E293B",
}

# ── Colonnes du CSV ──
COLONNES = [
    "NIP", "NDA", "NOM", "PRENOM", "DATE NAIS", "DATEVALEUR",
    "QUESTIONNAIRE", "Responsable", "Métier", "QUESTION", "REPONSE",
]


class AppQuestionnaires(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Analyse de Complétude — CH Cahors")
        self.geometry("1150x750")
        self.minsize(1000, 650)
        
        # Forcer le mode sombre
        ctk.set_appearance_mode("dark")
        self.configure(fg_color=THEME["bg"])

        self.csv_path: str | None = None
        self.donnees: list[dict] | None = None
        self.questionnaires_decouverts: dict[str, list[str]] = {}
        self.config: dict[str, list[str]] = {}
        self.checkbox_vars: dict[tuple[str, str], ctk.BooleanVar] = {}

        # ── État du Contrôle Qualité ──
        self.qc_service_var = ctk.StringVar(value="")
        self.qc_checkbox_vars: dict[str, ctk.BooleanVar] = {}
        self.qc_nb_sample_var = ctk.StringVar(value="10")

        self._build_ui()
        self._load_config()

    def _load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.config = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.config = {}

    def _save_config(self):
        config = self._get_current_selection()
        os.makedirs(os.path.dirname(os.path.abspath(CONFIG_FILE)), exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
        self.config = config
        messagebox.showinfo(
            "Sauvegarde",
            f"Configuration sauvegardée !\n\n"
            f"{sum(len(v) for v in config.values())} question(s) obligatoire(s) "
            f"pour {len(config)} questionnaire(s).",
        )

    def _build_ui(self):
        # Grille principale : Sidebar à gauche, Contenu à droite
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── 1. SIDEBAR (Colonne de gauche) ──
        self.sidebar = ctk.CTkFrame(
            self, 
            fg_color=THEME["bg_sidebar"], 
            corner_radius=0,
            border_width=1,
            border_color=THEME["border"]
        )
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(5, weight=1) # Espace flexible

        titre_lbl = ctk.CTkLabel(
            self.sidebar,
            text="🏥 CH Cahors\nAnalyse Complétude",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=THEME["text_title"],
            justify="center"
        )
        titre_lbl.grid(row=0, column=0, padx=20, pady=(35, 20))

        sep1 = ctk.CTkFrame(self.sidebar, height=1, fg_color=THEME["border"])
        sep1.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 25))

        self.btn_load = ctk.CTkButton(
            self.sidebar,
            text="📁  Charger le CSV",
            command=self._load_csv,
            fg_color=THEME["btn_load"],
            hover_color=THEME["btn_load_hover"],
            text_color=THEME["text_light"],
            font=ctk.CTkFont(size=14, weight="bold"),
            height=45,
            corner_radius=8
        )
        self.btn_load.grid(row=2, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.lbl_file = ctk.CTkLabel(
            self.sidebar,
            text="Aucun fichier",
            font=ctk.CTkFont(size=12),
            text_color=THEME["text_dim"],
            wraplength=180
        )
        self.lbl_file.grid(row=3, column=0, padx=20, pady=(0, 25))

        self.btn_save = ctk.CTkButton(
            self.sidebar,
            text="💾  Sauvegarder",
            command=self._save_config,
            fg_color=THEME["btn_save"],
            text_color=THEME["text_light"],
            hover_color=THEME["btn_save_hover"],
            font=ctk.CTkFont(size=13, weight="bold"),
            height=40,
            corner_radius=8,
            state="disabled"
        )
        self.btn_save.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="ew")

        # Résumé (en bas)
        self.summary_frame = ctk.CTkFrame(
            self.sidebar,
            fg_color=THEME["bg_section"],
            corner_radius=8
        )
        self.summary_frame.grid(row=6, column=0, padx=20, pady=(0, 20), sticky="ew")
        
        self.lbl_summary = ctk.CTkLabel(
            self.summary_frame,
            text="En attente de fichier...",
            font=ctk.CTkFont(size=13),
            text_color=THEME["text_dim"],
            justify="center",
            wraplength=180
        )
        self.lbl_summary.pack(padx=15, pady=20)

        self.btn_analyze = ctk.CTkButton(
            self.sidebar,
            text="📊  Générer le PDF",
            command=self._run_analysis,
            fg_color=THEME["btn_pdf"],
            hover_color=THEME["btn_pdf_hover"],
            text_color=THEME["text_light"],
            font=ctk.CTkFont(size=15, weight="bold"),
            height=50,
            corner_radius=8,
            state="disabled"
        )
        self.btn_analyze.grid(row=7, column=0, padx=20, pady=(0, 30), sticky="ew")

        # ── 2. ZONE PRINCIPALE AVEC TABVIEW (Colonne de droite) ──
        self.main_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_area.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.main_area.grid_rowconfigure(0, weight=1)
        self.main_area.grid_columnconfigure(0, weight=1)

        # TabView pour les deux modes
        self.tabview = ctk.CTkTabview(
            self.main_area,
            fg_color=THEME["bg_card"],
            segmented_button_fg_color=THEME["tab_unselected"],
            segmented_button_selected_color=THEME["tab_selected"],
            segmented_button_selected_hover_color=THEME["btn_load_hover"],
            segmented_button_unselected_color=THEME["tab_unselected"],
            segmented_button_unselected_hover_color=THEME["bg_section"],
            text_color=THEME["text_light"],
            corner_radius=12,
            border_width=1,
            border_color=THEME["border"],
        )
        self.tabview.grid(row=0, column=0, sticky="nsew")

        # Onglet 1 : Configuration
        self.tab_config = self.tabview.add("⚙️  Configuration Obligatoires")
        self.tab_config.grid_rowconfigure(1, weight=1)
        self.tab_config.grid_columnconfigure(0, weight=1)

        # En-tête de l'onglet Configuration
        header_config = ctk.CTkFrame(self.tab_config, fg_color="transparent")
        header_config.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        
        ctk.CTkLabel(
            header_config,
            text="Configuration des questions obligatoires",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=THEME["text_title"],
        ).pack(side="left")

        # Zone scrollable pour les questionnaires
        self.scroll_frame = ctk.CTkScrollableFrame(
            self.tab_config,
            fg_color=THEME["bg_card"],
            corner_radius=10,
            border_width=0,
        )
        self.scroll_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.scroll_frame.grid_columnconfigure(0, weight=1)

        self.lbl_placeholder = ctk.CTkLabel(
            self.scroll_frame,
            text="👈 Utilisez le bouton bleu 'Charger le CSV' à gauche",
            font=ctk.CTkFont(size=16),
            text_color=THEME["text_dim"]
        )
        self.lbl_placeholder.grid(row=0, column=0, pady=150)

        # ── Onglet 2 : Contrôle Qualité ──
        self.tab_qc = self.tabview.add("🔍  Contrôle Qualité")
        self.tab_qc.grid_rowconfigure(1, weight=1)
        self.tab_qc.grid_columnconfigure(0, weight=1)

        self._build_qc_tab()

    # ──────────────────────────────────────────────
    #  ONGLET CONTRÔLE QUALITÉ
    # ──────────────────────────────────────────────

    def _build_qc_tab(self):
        """Construit l'interface de l'onglet Contrôle Qualité."""

        # ── Barre supérieure : sélection du service + nb d'échantillons ──
        top_bar = ctk.CTkFrame(self.tab_qc, fg_color=THEME["bg_section"], corner_radius=10)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        top_bar.grid_columnconfigure(1, weight=1)

        # Label Service
        ctk.CTkLabel(
            top_bar,
            text="📋 Service :",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=THEME["text_title"],
        ).grid(row=0, column=0, padx=(15, 5), pady=12)

        # Menu déroulant des services
        self.qc_service_menu = ctk.CTkOptionMenu(
            top_bar,
            variable=self.qc_service_var,
            values=["— Charger un CSV d'abord —"],
            command=self._on_qc_service_changed,
            fg_color=THEME["btn_load"],
            button_color=THEME["btn_load_hover"],
            button_hover_color=THEME["btn_load"],
            text_color=THEME["text_light"],
            font=ctk.CTkFont(size=13),
            dropdown_fg_color=THEME["bg_sidebar"],
            dropdown_text_color=THEME["text_main"],
            dropdown_hover_color=THEME["bg_section"],
            width=350,
            height=35,
            corner_radius=8,
        )
        self.qc_service_menu.grid(row=0, column=1, padx=10, pady=12, sticky="w")

        # Nombre d'échantillons
        ctk.CTkLabel(
            top_bar,
            text="🎲 Échantillon :",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=THEME["text_title"],
        ).grid(row=0, column=2, padx=(20, 5), pady=12)

        self.qc_sample_entry = ctk.CTkEntry(
            top_bar,
            textvariable=self.qc_nb_sample_var,
            width=60,
            height=35,
            font=ctk.CTkFont(size=14),
            fg_color=THEME["bg_card"],
            text_color=THEME["text_main"],
            border_color=THEME["border"],
            corner_radius=8,
            justify="center",
        )
        self.qc_sample_entry.grid(row=0, column=3, padx=(0, 15), pady=12)

        # Bouton Export Excel
        self.btn_export_excel = ctk.CTkButton(
            top_bar,
            text="📥  Exporter Excel",
            command=self._export_qc_excel,
            fg_color=THEME["btn_excel"],
            hover_color=THEME["btn_excel_hover"],
            text_color=THEME["text_light"],
            font=ctk.CTkFont(size=14, weight="bold"),
            height=38,
            corner_radius=8,
            state="disabled",
        )
        self.btn_export_excel.grid(row=0, column=4, padx=(10, 15), pady=12)

        # ── Zone scrollable des questions pour le service choisi ──
        self.qc_scroll_frame = ctk.CTkScrollableFrame(
            self.tab_qc,
            fg_color=THEME["bg_card"],
            corner_radius=10,
            border_width=0,
        )
        self.qc_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.qc_scroll_frame.grid_columnconfigure(0, weight=1)

        self.qc_placeholder = ctk.CTkLabel(
            self.qc_scroll_frame,
            text="👆 Sélectionnez un service pour voir les questions",
            font=ctk.CTkFont(size=16),
            text_color=THEME["text_dim"],
        )
        self.qc_placeholder.grid(row=0, column=0, pady=150)

    def _on_qc_service_changed(self, service_name: str):
        """Appelé quand on change de service dans l'onglet QC."""
        if not self.donnees or service_name == "— Charger un CSV d'abord —":
            return

        # Nettoyer
        for w in self.qc_scroll_frame.winfo_children():
            w.destroy()
        self.qc_checkbox_vars.clear()

        # Trouver toutes les questions de ce service
        questions_du_service = self.questionnaires_decouverts.get(service_name, [])

        if not questions_du_service:
            ctk.CTkLabel(
                self.qc_scroll_frame,
                text="Aucune question trouvée pour ce service.",
                font=ctk.CTkFont(size=14),
                text_color=THEME["text_dim"],
            ).grid(row=0, column=0, pady=100)
            self.btn_export_excel.configure(state="disabled")
            return

        # Compter le nb de réponses NON VIDES pour chaque question
        comptage: dict[str, dict] = {}
        for row in self.donnees:
            if row.get("QUESTIONNAIRE", "").strip() != service_name:
                continue
            question = row.get("QUESTION", "").strip()
            reponse = row.get("REPONSE", "").strip()
            if question not in comptage:
                comptage[question] = {"total": 0, "remplies": 0}
            comptage[question]["total"] += 1
            if reponse:
                comptage[question]["remplies"] += 1

        # En-tête
        hdr = ctk.CTkFrame(
            self.qc_scroll_frame,
            fg_color=THEME["header_card"],
            corner_radius=8,
        )
        hdr.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            hdr,
            text=f"  {service_name}  —  {len(questions_du_service)} question(s) disponible(s)",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=THEME["text_light"],
            anchor="w",
        ).grid(row=0, column=0, padx=10, pady=10, sticky="w")

        # Boutons Tout / Aucun
        btn_box = ctk.CTkFrame(hdr, fg_color="transparent")
        btn_box.grid(row=0, column=1, padx=10, pady=5)

        ctk.CTkButton(
            btn_box, text="Tout", width=60, height=28,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=THEME["checkbox"],
            hover_color=THEME["checkbox_hover"],
            text_color=THEME["text_light"],
            corner_radius=6,
            command=lambda: self._toggle_all_qc(True),
        ).grid(row=0, column=0, padx=4)

        ctk.CTkButton(
            btn_box, text="Aucun", width=60, height=28,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=THEME["bg_sidebar"],
            hover_color=THEME["border"],
            text_color=THEME["text_main"],
            corner_radius=6,
            command=lambda: self._toggle_all_qc(False),
        ).grid(row=0, column=1, padx=4)

        # Liste des questions avec comptage
        q_container = ctk.CTkFrame(self.qc_scroll_frame, fg_color="transparent")
        q_container.grid(row=1, column=0, sticky="ew", padx=15, pady=(5, 10))
        q_container.grid_columnconfigure(0, weight=1)

        for question in questions_du_service:
            info = comptage.get(question, {"total": 0, "remplies": 0})
            nb_total = info["total"]
            nb_remplies = info["remplies"]

            var = ctk.BooleanVar(value=False)
            self.qc_checkbox_vars[question] = var

            row_frame = ctk.CTkFrame(q_container, fg_color="transparent")
            row_frame.pack(fill="x", pady=4)
            row_frame.grid_columnconfigure(0, weight=1)

            ctk.CTkCheckBox(
                row_frame,
                text=question,
                variable=var,
                font=ctk.CTkFont(size=13),
                text_color=THEME["text_main"],
                fg_color=THEME["checkbox"],
                hover_color=THEME["checkbox_hover"],
                checkmark_color="white",
                border_color=THEME["border"],
                border_width=2,
            ).grid(row=0, column=0, sticky="w")

            # Badge avec le nombre de réponses
            badge_text = f"{nb_remplies} rép."
            if nb_remplies >= 10:
                badge_color = THEME["success"]
            elif nb_remplies >= 5:
                badge_color = THEME["warning"]
            else:
                badge_color = THEME["danger"]

            badge = ctk.CTkLabel(
                row_frame,
                text=f"  {badge_text}  ({nb_total} total)  ",
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=THEME["text_light"],
                fg_color=badge_color,
                corner_radius=12,
                height=24,
            )
            badge.grid(row=0, column=1, padx=(10, 0), sticky="e")

        self.btn_export_excel.configure(state="normal")

    def _toggle_all_qc(self, state: bool):
        """Coche ou décoche toutes les questions dans l'onglet QC."""
        for var in self.qc_checkbox_vars.values():
            var.set(state)

    def _export_qc_excel(self):
        """Exporte les lignes tirées aléatoirement dans un fichier Excel formaté."""
        if not self.donnees:
            messagebox.showwarning("Attention", "Veuillez d'abord charger un fichier CSV.")
            return

        service = self.qc_service_var.get()
        if not service or service == "— Charger un CSV d'abord —":
            messagebox.showwarning("Attention", "Veuillez sélectionner un service.")
            return

        # Questions sélectionnées
        questions_selectionnees = [
            q for q, var in self.qc_checkbox_vars.items() if var.get()
        ]
        if not questions_selectionnees:
            messagebox.showwarning(
                "Attention",
                "Aucune question sélectionnée !\n"
                "Cochez au moins une question pour l'export.",
            )
            return

        # Nombre d'échantillons
        try:
            nb_sample = int(self.qc_nb_sample_var.get())
            if nb_sample < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror(
                "Erreur",
                "Le nombre d'échantillons doit être un entier positif.",
            )
            return

        # Construire les données pour chaque question
        tirage_par_question: dict[str, list[dict]] = {}
        for question in questions_selectionnees:
            # Filtrer les lignes : bon service, bonne question, réponse NON VIDE
            lignes_candidates = [
                row for row in self.donnees
                if row.get("QUESTIONNAIRE", "").strip() == service
                and row.get("QUESTION", "").strip() == question
                and row.get("REPONSE", "").strip() != ""
            ]

            # Tirage aléatoire
            if len(lignes_candidates) <= nb_sample:
                echantillon = lignes_candidates[:]
            else:
                echantillon = random.sample(lignes_candidates, nb_sample)

            tirage_par_question[question] = echantillon

        # Demander où sauvegarder
        output_path = filedialog.asksaveasfilename(
            title="Enregistrer l'export Excel",
            defaultextension=".xlsx",
            initialfile=f"controle_qualite_{service.replace(' ', '_')}.xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
        )
        if not output_path:
            return

        # Générer l'Excel
        try:
            self._generer_excel(service, tirage_par_question, nb_sample, output_path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Erreur lors de la génération :\n{exc}")
            return

        # Résumé
        total_lignes = sum(len(v) for v in tirage_par_question.values())
        messagebox.showinfo(
            "Export réussi ✅",
            f"Fichier Excel généré !\n\n"
            f"Service : {service}\n"
            f"Questions : {len(questions_selectionnees)}\n"
            f"Lignes exportées : {total_lignes}\n"
            f"Échantillon demandé : {nb_sample} par question",
        )

        # Ouvrir le fichier
        try:
            os.startfile(os.path.abspath(output_path))
        except Exception:
            pass

    def _generer_excel(
        self,
        service: str,
        tirage_par_question: dict[str, list[dict]],
        nb_sample: int,
        chemin_sortie: str,
    ):
        """Génère un fichier Excel formaté avec openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Contrôle Qualité"

        # ── Styles ──
        fill_header_service = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        fill_header_question = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        fill_header_cols = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        fill_alt_1 = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
        fill_alt_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        font_service = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        font_question = Font(name="Calibri", bold=True, size=13, color="E0E7FF")
        font_header_col = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        font_data = Font(name="Calibri", size=10, color="2C3E50")
        font_info = Font(name="Calibri", italic=True, size=10, color="6B7280")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin", color="DEE2E6"),
            right=Side(style="thin", color="DEE2E6"),
            top=Side(style="thin", color="DEE2E6"),
            bottom=Side(style="thin", color="DEE2E6"),
        )

        # Largeurs de colonnes
        col_widths = {
            "A": 14, "B": 16, "C": 16, "D": 14, "E": 14,
            "F": 22, "G": 30, "H": 22, "I": 12, "J": 35, "K": 45,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        current_row = 1

        # ── Ligne titre du service ──
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"🏥 Contrôle Qualité — {service}"
        cell.font = font_service
        cell.fill = fill_header_service
        cell.alignment = align_center
        ws.row_dimensions[current_row].height = 40
        current_row += 1

        # Ligne date + info
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        cell = ws.cell(row=current_row, column=1)
        date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
        cell.value = f"Généré le {date_str} — Échantillon de {nb_sample} réponse(s) par question"
        cell.font = font_info
        cell.alignment = align_center
        ws.row_dimensions[current_row].height = 22
        current_row += 2

        # ── Pour chaque question sélectionnée ──
        for question, lignes in tirage_par_question.items():
            # Sous-titre question
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            cell = ws.cell(row=current_row, column=1)
            nb_remplies_total = sum(
                1 for row in self.donnees
                if row.get("QUESTIONNAIRE", "").strip() == service
                and row.get("QUESTION", "").strip() == question
                and row.get("REPONSE", "").strip() != ""
            )
            cell.value = f"📝 {question}  —  {len(lignes)} tirée(s) / {nb_remplies_total} réponse(s) remplie(s)"
            cell.font = font_question
            cell.fill = fill_header_question
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 32
            current_row += 1

            if not lignes:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
                cell = ws.cell(row=current_row, column=1)
                cell.value = "Aucune réponse remplie trouvée pour cette question."
                cell.font = font_info
                cell.alignment = align_center
                current_row += 2
                continue

            # En-tête des colonnes
            for col_idx, header_name in enumerate(COLONNES, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header_name
                cell.font = font_header_col
                cell.fill = fill_header_cols
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 26
            current_row += 1

            # Lignes de données
            for data_idx, row_data in enumerate(lignes):
                fill = fill_alt_1 if data_idx % 2 == 0 else fill_alt_2
                for col_idx, col_name in enumerate(COLONNES, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = row_data.get(col_name, "")
                    cell.font = font_data
                    cell.fill = fill
                    cell.alignment = align_left
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            current_row += 1  # Espace entre les questions

        # Figer la première ligne
        ws.freeze_panes = "A4"

        wb.save(chemin_sortie)

    # ──────────────────────────────────────────────
    #  ONGLET CONFIGURATION (existant)
    # ──────────────────────────────────────────────

    def _load_csv(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier CSV",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous", "*.*")],
        )
        if not path:
            return

        self.csv_path = path
        nom_fichier = os.path.basename(path)
        self.lbl_file.configure(
            text=f"📄 {nom_fichier}",
            text_color=THEME["btn_load"],
            font=ctk.CTkFont(size=12, weight="bold")
        )

        try:
            self.donnees = lire_csv(path)
        except Exception as exc:
            messagebox.showerror("Erreur de lecture", str(exc))
            return

        decouverts: dict[str, set[str]] = {}
        for row in self.donnees:
            q = row.get("QUESTIONNAIRE", "").strip()
            question = row.get("QUESTION", "").strip()
            if q and question:
                decouverts.setdefault(q, set()).add(question)

        self.questionnaires_decouverts = {
            q: sorted(questions) for q, questions in sorted(decouverts.items())
        }

        self._build_questionnaire_list()

        # Mettre à jour le menu déroulant QC
        services_list = sorted(self.questionnaires_decouverts.keys())
        self.qc_service_menu.configure(values=services_list)
        if services_list:
            self.qc_service_var.set(services_list[0])
            self._on_qc_service_changed(services_list[0])

        self.btn_save.configure(state="normal")
        self.btn_analyze.configure(state="normal")

        nb_q = len(self.questionnaires_decouverts)
        self.lbl_summary.configure(
            text=f"✅ Prêt !\n\n{nb_q} services détectés.",
            text_color=THEME["success"],
            font=ctk.CTkFont(size=14, weight="bold")
        )

    def _build_questionnaire_list(self):
        for w in self.scroll_frame.winfo_children():
            w.destroy()
        self.checkbox_vars.clear()

        row_idx = 0

        # Calcul du comptage pour tous les questionnaires
        comptage: dict[tuple[str, str], dict] = {}
        if self.donnees:
            for row in self.donnees:
                questionnaire = row.get("QUESTIONNAIRE", "").strip()
                question = row.get("QUESTION", "").strip()
                reponse = row.get("REPONSE", "").strip()
                
                cle = (questionnaire, question)
                if cle not in comptage:
                    comptage[cle] = {"total": 0, "remplies": 0}
                
                comptage[cle]["total"] += 1
                if reponse:
                    comptage[cle]["remplies"] += 1

        for questionnaire in sorted(self.questionnaires_decouverts):
            questions = self.questionnaires_decouverts[questionnaire]
            config_questions = self.config.get(questionnaire, [])

            # En-tête du questionnaire
            hdr = ctk.CTkFrame(
                self.scroll_frame,
                fg_color=THEME["header_card"],
                corner_radius=8,
            )
            hdr.grid(row=row_idx, column=0, sticky="ew", padx=15, pady=(20, 8))
            hdr.grid_columnconfigure(0, weight=1)

            nb_obligatoires = sum(1 for q in questions if q in config_questions)
            
            ctk.CTkLabel(
                hdr,
                text=f"  {questionnaire}  ({nb_obligatoires}/{len(questions)} obligatoires)",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=THEME["text_light"],
                anchor="w",
            ).grid(row=0, column=0, padx=10, pady=10, sticky="w")

            # Boutons de sélection rapide
            btn_box = ctk.CTkFrame(hdr, fg_color="transparent")
            btn_box.grid(row=0, column=1, padx=10, pady=5)

            ctk.CTkButton(
                btn_box, text="Tout", width=60, height=28,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color=THEME["checkbox"],
                hover_color=THEME["checkbox_hover"],
                text_color=THEME["text_light"],
                corner_radius=6,
                command=lambda q=questionnaire: self._toggle_all(q, True),
            ).grid(row=0, column=0, padx=4)

            ctk.CTkButton(
                btn_box, text="Aucun", width=60, height=28,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color=THEME["bg_sidebar"],
                hover_color=THEME["border"],
                text_color=THEME["text_main"],
                corner_radius=6,
                command=lambda q=questionnaire: self._toggle_all(q, False),
            ).grid(row=0, column=1, padx=4)

            row_idx += 1

            # Liste des questions
            q_container = ctk.CTkFrame(self.scroll_frame, fg_color="transparent")
            q_container.grid(row=row_idx, column=0, sticky="ew", padx=25, pady=0)
            q_container.grid_columnconfigure(0, weight=1)

            for question in questions:
                is_checked = question in config_questions
                var = ctk.BooleanVar(value=is_checked)
                self.checkbox_vars[(questionnaire, question)] = var

                row_frame = ctk.CTkFrame(q_container, fg_color="transparent")
                row_frame.pack(fill="x", pady=4)
                row_frame.grid_columnconfigure(0, weight=1)

                ctk.CTkCheckBox(
                    row_frame,
                    text=question,
                    variable=var,
                    font=ctk.CTkFont(size=13),
                    text_color=THEME["text_main"],
                    fg_color=THEME["checkbox"],
                    hover_color=THEME["checkbox_hover"],
                    checkmark_color="white",
                    border_color=THEME["border"],
                    border_width=2,
                ).grid(row=0, column=0, sticky="w")

                # Badge avec le nombre de réponses
                info = comptage.get((questionnaire, question), {"total": 0, "remplies": 0})
                nb_total = info["total"]
                nb_remplies = info["remplies"]
                
                taux = (nb_remplies / nb_total * 100) if nb_total > 0 else 0
                
                if taux >= 100:
                    badge_color = THEME["success"]
                elif taux >= 60:
                    badge_color = THEME["warning"]
                else:
                    badge_color = THEME["danger"]
                    
                badge_text = f"{nb_remplies} rép."
                
                badge = ctk.CTkLabel(
                    row_frame,
                    text=f"  {badge_text}  ({nb_total} total)  ",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    text_color=THEME["text_light"],
                    fg_color=badge_color,
                    corner_radius=12,
                    height=24,
                )
                badge.grid(row=0, column=1, padx=(10, 0), sticky="e")

            row_idx += 1

    def _toggle_all(self, questionnaire: str, state: bool):
        for (q, _), var in self.checkbox_vars.items():
            if q == questionnaire:
                var.set(state)

    def _get_current_selection(self) -> dict[str, list[str]]:
        config: dict[str, list[str]] = {}
        for (questionnaire, question), var in self.checkbox_vars.items():
            if var.get():
                config.setdefault(questionnaire, []).append(question)
        return config

    def _run_analysis(self):
        if not self.donnees:
            messagebox.showwarning("Attention", "Veuillez d'abord charger un fichier CSV.")
            return

        config = self._get_current_selection()

        if not config:
            messagebox.showwarning(
                "Attention",
                "Aucune question obligatoire sélectionnée !\n"
                "Cochez au moins une question pour lancer l'analyse.",
            )
            return

        stats = analyser_completude(self.donnees, config)

        if not stats:
            messagebox.showinfo(
                "Résultat",
                "Aucune donnée correspondante trouvée.\n"
                "Vérifiez que les questions cochées sont présentes dans le CSV.",
            )
            return

        output_path = filedialog.asksaveasfilename(
            title="Enregistrer le rapport PDF",
            defaultextension=".pdf",
            initialfile="rapport_completude.pdf",
            filetypes=[("Fichier PDF", "*.pdf")],
        )
        if not output_path:
            return

        questions_analysees = list(set(q for _, q in stats.keys()))

        try:
            generer_rapport_pdf(self.donnees, stats, questions_analysees, output_path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Erreur lors de la génération :\n{exc}")
            return

        total = sum(d["total"] for d in stats.values())
        vides = sum(d["vides"] for d in stats.values())
        taux = ((total - vides) / total * 100) if total > 0 else 0
        nb_q = len(set(q for q, _ in stats.keys()))

        if taux >= 95:
            color = THEME["success"]
        elif taux >= 85:
            color = THEME["warning"]
        else:
            color = THEME["danger"]

        self.lbl_summary.configure(
            text=f"Rapport généré !\n\n"
                 f"{nb_q} services\n"
                 f"Taux : {taux:.1f}%\n"
                 f"{vides} vides / {total}",
            text_color=color,
            font=ctk.CTkFont(size=14, weight="bold")
        )

        try:
            os.startfile(os.path.abspath(output_path))
        except Exception:
            pass

if __name__ == "__main__":
    app = AppQuestionnaires()
    app.mainloop()
